from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import time
from selenium.webdriver.common.by import By
from openpyxl import load_workbook




wb = load_workbook("Contacts.xlsx")
wb.active = 0
sheet = wb.active
driver =webdriver.Chrome("C:/Users/Desktop/Desktop/WhatsApp_Automation/chromedriver.exe")
driver.get("https://web.whatsapp.com/")
n = input()
#reading specific column
for i in range(1,5):
    name = sheet.cell(row=i, column=1).value
    phone_Number = sheet.cell(row=i, column=2).value

    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')

    msg = "Hi" + name +" , you are Invited to demo"
    driver.get("https://api.whatsapp.com/send?phone=91" + str(phone_Number) + "&text=" + str(msg))
    print("https://api.whatsapp.com/send?phone=91" + str(phone_Number) + "&text=" + str(msg))
    if(i==1):
        time.sleep(3)
        continueToChat = driver.find_element_by_id("action-button")
        action = ActionChains(driver)
        action.click(continueToChat)
        action.perform()
        time.sleep(1)
        continueWithWebBrowser = driver.find_elements(By.CLASS_NAME, "_36or")[3]
        action = ActionChains(driver)
        action.click(continueWithWebBrowser)
        action.perform()
    time.sleep(6)
    sendMsg = driver.find_elements_by_class_name("_4sWnG")
    print(sendMsg)
    action = ActionChains(driver)
    action.click(sendMsg[0])
    action.perform()
























  


